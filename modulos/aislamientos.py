import streamlit as st
import pandas as pd
import numpy as np
import time
from io import BytesIO
from datetime import datetime, timedelta
from streamlit_gsheets import GSheetsConnection

# Librerías para Excel Profesional
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Librerías para PDF Profesional
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# --- CONFIGURACIÓN DE PÁGINA ---
st.set_page_config(page_title="Control de Aislamientos", page_icon="🦠", layout="wide")

# URLs de Google Sheets
SHEET_URL_ORIGEN = "https://docs.google.com/spreadsheets/d/e/2PACX-1vRmU8ibxYHge7Mq0bcdBz5oa7TPtWt6-3uxungBZlfHCM7oUzUy2TNL43tOmeHOzHebX-xGfvqFcxiy/pub?gid=1090111501&single=true&output=csv"
SHEET_URL_EDITABLE = "https://docs.google.com/spreadsheets/d/1LfQTTfto_I5bpLIyiWblfypD3gu99MoWldmW9bmuJ4A/edit"

# --- CONEXIÓN ---
conn = st.connection("gsheets", type=GSheetsConnection)

# --- FUNCIONES DE FORMATO DE REPORTES ---

def aplicar_formato_excel_oficial(writer, sheet_name, df, titulo_reporte):
    ws = writer.sheets[sheet_name]
    hoy = datetime.now()
    vencimiento = hoy + timedelta(days=7)
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    num_cols = len(df.columns)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    titulo_texto = f"{titulo_reporte} DEL {hoy.strftime('%d/%m/%Y')} AL {vencimiento.strftime('%d/%m/%Y')} (PARA LOS 3 TURNOS Y FINES DE SEMANA)"
    
    cell_h = ws.cell(row=1, column=1)
    cell_h.value = titulo_texto
    cell_h.alignment = center_align
    cell_h.font = Font(bold=True, size=11)

    for col_num, value in enumerate(df.columns, 1):
        cell = ws.cell(row=2, column=col_num, value=value)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    for row in ws.iter_rows(min_row=3, max_row=len(df)+2, min_col=1, max_col=num_cols):
        for cell in row:
            cell.border = border
            cell.alignment = center_align

    fila_pie = len(df) + 3
    ws.merge_cells(start_row=fila_pie, start_column=1, end_row=fila_pie, end_column=num_cols)
    leyenda = "Comentario: de acuerdo con la Norma Oficial Mexicana NOM-045-SSA2-2005, Para la vigilancia epidemiológica, prevención y control de las infecciones nosocomiales. NINGUN RECIPIENTE QUE CONTENGA EL INSUMO DEBERÁ SER RELLENADO O REUTILIZADO."
    
    cell_nom = ws.cell(row=fila_pie, column=1, value=leyenda)
    cell_nom.alignment = center_align
    cell_nom.font = Font(size=9, italic=True)
    ws.row_dimensions[fila_pie].height = 45

    fila_firma = fila_pie + 1
    ws.merge_cells(start_row=fila_firma, start_column=1, end_row=fila_firma, end_column=num_cols)
    cell_auth = ws.cell(row=fila_firma, column=1, value="AUTORIZÓ: DRA. BRENDA CASTILLO MATUS")
    cell_auth.alignment = center_align
    cell_auth.font = Font(bold=True, size=11)

    for i in range(1, num_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 25

def generar_pdf_oficial(df):
    output = BytesIO()
    doc = SimpleDocTemplate(
        output, 
        pagesize=landscape(letter), 
        topMargin=20, 
        bottomMargin=20, 
        leftMargin=30, 
        rightMargin=30
    )
    styles = getSampleStyleSheet()
    
    num_pacientes = len(df)
    font_size_base = 7 if num_pacientes < 20 else 6
    leading_base = 8 if num_pacientes < 20 else 7
    
    estilo_titulo = ParagraphStyle('T', parent=styles['Heading2'], alignment=1, fontSize=11, spaceAfter=2)
    estilo_subtitulo = ParagraphStyle('S', parent=styles['Normal'], alignment=1, fontSize=9, spaceAfter=10)
    estilo_celda = ParagraphStyle('cell', parent=styles['Normal'], fontSize=font_size_base, alignment=1, leading=leading_base)
    estilo_encabezado = ParagraphStyle('header', parent=styles['Normal'], fontSize=font_size_base + 1, textColor=colors.whitesmoke, alignment=1, fontName='Helvetica-Bold')
    estilo_leyenda = ParagraphStyle('footer', parent=styles['Normal'], fontSize=7, italic=True, alignment=1, leading=8)
    
    elements = []
    hoy = datetime.now()
    vencimiento = hoy + timedelta(days=7)
    
    elements.append(Paragraph("CENSO DE AISLAMIENTOS", estilo_titulo))
    elements.append(Paragraph(
        f"VIGENCIA: DEL {hoy.strftime('%d/%m/%Y')} AL {vencimiento.strftime('%d/%m/%Y')} (PARA LOS 3 TURNOS Y FINES DE SEMANA)", 
        estilo_subtitulo
    ))
    
    data = [[Paragraph(col, estilo_encabezado) for col in df.columns]]
    for row in df.values:
        data.append([Paragraph(str(item), estilo_celda) for item in row])
    
    col_widths = [50, 65, 210, 140, 90, 140]
    
    t = RLTable(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ]))
    
    t.hAlign = 'CENTER' 
    elements.append(t)
    elements.append(Spacer(1, 10))
    
    leyenda = "Comentario: de acuerdo con la Norma Oficial Mexicana NOM-045-SSA2-2005, Para la vigilancia epidemiológica, prevención y control de las infecciones nosocomiales. NINGUN RECIPIENTE QUE CONTENGA EL INSUMO DEBERÁ SER RELLENADO O REUTILIZADO."
    elements.append(Paragraph(leyenda, estilo_leyenda))
    elements.append(Spacer(1, 5))
    elements.append(Paragraph("<b>AUTORIZÓ: DRA. BRENDA CASTILLO MATUS</b>", estilo_subtitulo))

    doc.build(elements)
    return output.getvalue()

# --- LÓGICA DE DATOS ---

@st.cache_data(ttl=2)
def cargar_datos_aislamiento():
    try:
        url_final = f"{SHEET_URL_ORIGEN}&cachebust={time.time()}"
        df = pd.read_csv(url_final, skiprows=1, engine='python')
        df = df.iloc[:, 1:10]
        
        # Estandarizar columnas
        df.columns = [str(c).strip().replace('\n', ' ').upper() for c in df.columns]
        
        # Rellenar datos
        df["CAMA"] = df["CAMA"].ffill()
        df["NOMBRE"] = df["NOMBRE"].ffill()
        
        # --- FILTRO DE VIGENCIA (FECHA DE TÉRMINO VACÍA) ---
        if "FECHA DE TÉRMINO" in df.columns:
            # Convertimos a string y tratamos diversos tipos de nulos
            df["FECHA DE TÉRMINO"] = df["FECHA DE TÉRMINO"].astype(str).replace(['nan', 'None', ' ', '', 'NaT'], np.nan)
            df = df[df["FECHA DE TÉRMINO"].isna()].copy()
        
        # Consolidación de tipos de aislamiento
        def consolidar(group):
            res = group.iloc[0].copy()
            if "TIPO DE AISLAMIENTO" in group.columns:
                tipos = group["TIPO DE AISLAMIENTO"].dropna().unique()
                res["TIPO DE AISLAMIENTO"] = " / ".join(map(str, tipos)) if len(tipos) > 0 else np.nan
            return res

        if not df.empty:
            df = df.groupby(["CAMA", "NOMBRE"], as_index=False, sort=False).apply(consolidar)
            df = df.reset_index(drop=True)
        
        # Orden de columnas
        cols_orden = ["CAMA", "REGISTRO", "NOMBRE", "TIPO DE AISLAMIENTO", "FECHA DE INICIO"]
        df = df[[c for c in cols_orden if c in df.columns]].copy()
        df["INSUMO"] = "JABÓN/SANITAS"
        
        # Limpieza final
        df = df.replace(['nan', 'None', '', ' '], np.nan).dropna(subset=["CAMA", "NOMBRE"])
        return df.reset_index(drop=True)
    except Exception as e:
        st.error(f"Error al cargar datos: {e}")
        return pd.DataFrame()

# --- INTERFAZ ---
st.title("🦠 Gestión de Vigilancia Epidemiológica")

tab1, tab2 = st.tabs(["🔍 Monitor y Edición", "📝 Insumos Aislamientos"])

with tab1:
    df_actual = cargar_datos_aislamiento()
    
    if not df_actual.empty:
        # --- CÁLCULOS DE CONTEO ---
        total_general = len(df_actual)
        mask_protector = df_actual["TIPO DE AISLAMIENTO"].str.contains("PROTECTOR", case=False, na=False)
        df_protectores = df_actual[mask_protector]
        total_protectores = len(df_protectores)
        
        # --- RENDERIZADO DE MÉTRICAS ---
        m1, m2 = st.columns(2)
        with m1:
            st.metric("Total Pacientes Aislados", total_general)
        with m2:
            st.metric("Aislamientos Protectores", total_protectores)
        
        c1, c2, c3 = st.columns(3)
        with c1:
            if st.button("🔄 Actualizar Monitor", use_container_width=True):
                st.cache_data.clear()
                st.rerun()
        with c2:
            if st.button("🚀 Sincronizar Drive", use_container_width=True):
                conn.update(spreadsheet=SHEET_URL_EDITABLE, data=df_actual)
                st.success("Drive Actualizado")
        with c3:
            st.link_button("📂 Abrir Sheets", SHEET_URL_EDITABLE, use_container_width=True)

        st.divider()
        
        # Sección de Edición Principal
        st.subheader("📋 Censo General (Editable)")
        df_drive = conn.read(spreadsheet=SHEET_URL_EDITABLE, ttl=0)
        
        if not df_drive.empty:
            df_ed = st.data_editor(df_drive, use_container_width=True, num_rows="dynamic", hide_index=True)
            if st.button("💾 Guardar Cambios en Drive", use_container_width=True):
                conn.update(spreadsheet=SHEET_URL_EDITABLE, data=df_ed.reset_index(drop=True))
                st.toast("Datos guardados", icon="✅")
        else:
            st.warning("No se encontraron datos en la hoja de edición.")
        
        # Sección Separada para Protectores
        if total_protectores > 0:
            with st.expander("🛡️ Ver detalle de Aislamientos Protectores"):
                st.dataframe(df_protectores, use_container_width=True, hide_index=True)
    else:
        st.warning("No hay pacientes con aislamientos vigentes según la base de datos.")
        if st.button("🔄 Reintentar Carga"):
            st.cache_data.clear()
            st.rerun()

with tab2:
    st.header("Generación de Reportes de Insumos")
    
    if not df_actual.empty:
        st.info("Descarga el censo de insumos. El PDF está configurado para ajustarse a una sola página.")
        col_ex, col_pdf = st.columns(2)
        
        with col_ex:
            output_ex = BytesIO()
            with pd.ExcelWriter(output_ex, engine='openpyxl') as writer:
                df_actual.to_excel(writer, index=False, sheet_name="INSUMOS", startrow=1)
                aplicar_formato_excel_oficial(writer, "INSUMOS", df_actual, "INSUMOS AISLAMIENTOS")
            
            st.download_button(
                "💾 DESCARGAR EXCEL",
                output_ex.getvalue(),
                f"Insumos_Aislamiento_{datetime.now().strftime('%d%m%Y')}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, type="primary"
            )
            
        with col_pdf:
            pdf_data = generar_pdf_oficial(df_actual)
            st.download_button(
                "📄 DESCARGAR PDF (UNA SOLA HOJA)",
                pdf_data,
                f"Insumos_Aislamiento_{datetime.now().strftime('%d%m%Y')}.pdf",
                "application/pdf",
                use_container_width=True
            )
    else:
        st.error("No hay datos disponibles para generar reportes.")
