import streamlit as st
import pandas as pd
import re
from io import BytesIO
from datetime import datetime, timedelta
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Librerías para el PDF profesional
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# --- CONFIGURACIÓN ---
SERVICIOS_INSUMOS_FILTRO = [
    "HEMATOLOGIA", "HEMATOLOGIA PEDIATRICA", "ONCOLOGIA PEDIATRICA",
    "NEONATOLOGIA", "INFECTOLOGIA PEDIATRICA", "U.C.I.N.",
    "U.T.I.P.", "TERAPIA POSQUIRURGICA", "UNIDAD DE QUEMADOS",
    "ONCOLOGIA MEDICA", "UCIA"
]

# --- FUNCIONES DE FORMATO EXCEL ---

def aplicar_formato_oficial(writer, sheet_name, df, servicio_nombre):
    """Aplica formato profesional: encabezados azules, vigencia, centrado y firma."""
    ws = writer.sheets[sheet_name]
    hoy = datetime.now()
    vencimiento = hoy + timedelta(days=7)
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 1. Título de Vigencia centrado
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    titulo = f"{servicio_nombre} DEL {hoy.strftime('%d/%m/%Y')} AL {vencimiento.strftime('%d/%m/%Y')} (PARA LOS 3 TURNOS Y FINES DE SEMANA)"
    cell_h = ws.cell(row=1, column=1, value=titulo)
    cell_h.alignment = center_align
    cell_h.font = Font(bold=True, size=11)

    # 2. Encabezados
    for col_num, value in enumerate(df.columns, 1):
        cell = ws.cell(row=2, column=col_num, value=value)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # 3. Cuerpo de datos centrado
    for row in ws.iter_rows(min_row=3, max_row=len(df)+2, min_col=1, max_col=8):
        for cell in row:
            cell.border = border
            cell.alignment = center_align
    
    # Ancho de columnas
    for i in range(1, 9):
        ws.column_dimensions[get_column_letter(i)].width = 20

    # 4. Pie de Página (Leyenda NOM-045 y Firma)
    lr = ws.max_row
    ws.merge_cells(start_row=lr + 1, start_column=1, end_row=lr + 1, end_column=8)
    leyenda = ("Comentario: de acuerdo con la Norma Oficial Mexicana NOM-045-SSA2-2005, "
               "Para la vigilancia epidemiológica, prevención y control de las infecciones nosocomiales. "
               "NINGUN RECIPIENTE QUE CONTENGA EL INSUMO DEBERÁ SER RELLENADO O REUTILIZADO.")
    cell_nom = ws.cell(row=lr + 1, column=1, value=leyenda)
    cell_nom.alignment = center_align
    cell_nom.font = Font(size=9, italic=True)
    ws.row_dimensions[lr + 1].height = 45

    ws.merge_cells(start_row=lr + 2, start_column=1, end_row=lr + 2, end_column=8)
    cell_auth = ws.cell(row=lr + 2, column=1, value="AUTORIZÓ: DRA. BRENDA CASTILLO MATUS")
    cell_auth.alignment = center_align
    cell_auth.font = Font(bold=True)

# --- FUNCIONES DE FORMATO PDF ---

def generar_pdf_oficial(dict_especialidades):
    """Genera PDF centrado con leyenda completa de la NOM."""
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(letter), topMargin=30, bottomMargin=30, leftMargin=40, rightMargin=40)
    styles = getSampleStyleSheet()
    elements = []
    
    hoy = datetime.now()
    vencimiento = hoy + timedelta(days=7)
    f_rango = f"DEL {hoy.strftime('%d/%m/%Y')} AL {vencimiento.strftime('%d/%m/%Y')}"

    # Estilos de párrafo
    title_style = ParagraphStyle('T', parent=styles['Heading2'], alignment=1, fontSize=11, spaceAfter=10)
    footer_style = ParagraphStyle('F', parent=styles['Normal'], fontSize=8, italic=True, alignment=1, leading=10)
    auth_style = ParagraphStyle('A', parent=styles['Normal'], fontSize=10, bold=True, alignment=1, spaceBefore=10)

    for serv, df in dict_especialidades.items():
        # Título centrado
        elements.append(Paragraph(f"INSUMOS {serv} {f_rango}<br/>(PARA LOS 3 TURNOS Y FINES DE SEMANA)", title_style))
        
        # Tabla centrada
        data = [df.columns.tolist()] + df.values.tolist()
        col_widths = [45, 60, 180, 45, 40, 70, 110, 110]
        t = RLTable(data, repeatRows=1, colWidths=col_widths)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        t.hAlign = 'CENTER'
        elements.append(t)
        elements.append(Spacer(1, 15))
        
        # Leyenda completa corregida
        leyenda_nom = ("Comentario: de acuerdo con la Norma Oficial Mexicana NOM-045-SSA2-2005, "
                       "Para la vigilancia epidemiológica, prevención y control de las infecciones nosocomiales. "
                       "NINGUN RECIPIENTE QUE CONTENGA EL INSUMO DEBERÁ SER RELLENADO O REUTILIZADO.")
        elements.append(Paragraph(leyenda_nom, footer_style))
        
        # Firma centrada
        elements.append(Paragraph("<b>AUTORIZÓ: DRA. BRENDA CASTILLO MATUS</b>", auth_style))
        elements.append(PageBreak())

    doc.build(elements)
    return output.getvalue()

# --- LÓGICA DE PROCESAMIENTO ---

def obtener_especialidad_real(cama, esp_html):
    c = str(cama).strip().upper()
    esp_clean = esp_html.replace("ESPECIALIDAD:", "").replace("&NBSP;", "").strip().upper()
    if c.startswith("55"): return "U.C.I.N."
    if c.startswith("45"): return "NEONATOLOGIA" 
    if c.startswith("56"): return "U.T.I.P."
    if c.startswith("85"): return "UNIDAD DE QUEMADOS"
    if c.startswith("73"): return "UCIA"
    if c.isdigit() and 7401 <= int(c) <= 7409: return "TERAPIA POSQUIRURGICA"
    return esp_clean

# --- INTERFAZ STREAMLIT ---
st.title("📦 Censo de Insumos: Especialidades")

# Configuración de barra lateral para subir archivo
with st.sidebar:
    st.header("Carga de Datos")
    archivo = st.file_uploader("Sube el archivo HTML (Censo)", type=["html"])
    if archivo:
        st.session_state['archivo_compartido'] = archivo

if 'archivo_compartido' not in st.session_state:
    st.info("👈 Sube el archivo HTML en la barra lateral para comenzar.")
else:
    try:
        # Procesar HTML
        tablas = pd.read_html(st.session_state['archivo_compartido'])
        df_raw = max(tablas, key=len)
        col0_str = df_raw.iloc[:, 0].fillna("").astype(str).str.upper()
        
        datos_html = []
        dict_final = {}
        IGNORAR = ["PACIENTES", "TOTAL", "SUBTOTAL", "PÁGINA", "IMPRESIÓN", "1111"]

        esp_actual = ""
        for i, val in enumerate(col0_str):
            if "ESPECIALIDAD:" in val:
                esp_actual = val; continue
            
            fila = [str(x).strip() for x in df_raw.iloc[i].values]
            if any(x in fila[0] or x in fila[1] for x in IGNORAR): continue

            # Validar si es una fila de paciente (Registro con al menos 5 dígitos)
            if len(fila) > 1 and len(fila[1]) >= 5 and any(char.isdigit() for char in fila[1]):
                esp_real = obtener_especialidad_real(fila[0], esp_actual)
                if esp_real in SERVICIOS_INSUMOS_FILTRO:
                    datos_html.append({
                        "CAMA": fila[0], "REGISTRO": fila[1], "PACIENTE": fila[2],
                        "SEXO": fila[3], "EDAD": "".join(re.findall(r'\d+', fila[4])),
                        "FECHA DE INGRESO": fila[9], "TIPO DE PRECAUCIONES": "ESTÁNDAR",
                        "INSUMO": "JABÓN/SANITAS", "ESP_REAL": esp_real
                    })

        if datos_html:
            df_full = pd.DataFrame(datos_html)
            servicios_detectados = sorted(df_full["ESP_REAL"].unique())
            
            for serv in servicios_detectados:
                df_s = df_full[df_full["ESP_REAL"] == serv].drop(columns="ESP_REAL")
                dict_final[serv] = df_s
                with st.expander(f"📍 {serv} ({len(df_s)} pacientes)"):
                    st.table(df_s)

            st.divider()
            c1, c2 = st.columns(2)
            
            with c1:
                # Excel
                out_ex = BytesIO()
                with pd.ExcelWriter(out_ex, engine='openpyxl') as writer:
                    for serv, df in dict_final.items():
                        hoja = serv[:30].replace("/", "-")
                        df.to_excel(writer, index=False, sheet_name=hoja, startrow=1)
                        aplicar_formato_oficial(writer, hoja, df, f"INSUMOS {serv}")
                
                st.download_button("💾 DESCARGAR EXCEL", out_ex.getvalue(), 
                                   f"Insumos_{datetime.now().strftime('%d%m%Y')}.xlsx", 
                                   use_container_width=True, type="primary")

            with c2:
                # PDF
                pdf_bytes = generar_pdf_oficial(dict_final)
                st.download_button("📄 DESCARGAR PDF", pdf_bytes, 
                                   f"Insumos_{datetime.now().strftime('%d%m%Y')}.pdf", 
                                   "application/pdf", use_container_width=True)
        else:
            st.warning("No se detectaron pacientes de las especialidades seleccionadas.")

    except Exception as e:
        st.error(f"Error al procesar el archivo: {e}")
